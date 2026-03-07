import pandas
import openpyxl
from openpyxl import load_workbook
import datetime
from datetime import datetime, timedelta
import os


# Создание ссылки
name_of_tab = "1FiMov0r4UUDKT6A56NWMImpoUakDC2YDevgaOpJQ7Qc"
page = "70289591"
url = f"https://docs.google.com/spreadsheets/d/{name_of_tab}/export?format=csv&gid={page}"

# Запись файла и перевод его в формат xlsx
open("check_table.xlsx", "w").close()
pandas.read_csv(url).to_excel("check_table.xlsx", index=False)

# Русификация дней недели
days_ru = ['понедельник', 'вторник', 'среда',
           'четверг', 'пятница', 'суббота', 'воскресенье']

# Определение завтрашнего дня
tomorrow = datetime.now() + timedelta(days=1)
day_name = days_ru[tomorrow.weekday()]
tomorrow = tomorrow.strftime('%d.%m.%Y') + f' {day_name}'

# Определение послезавтра, проверка на воскресенье
nextmorrow = datetime.now() + timedelta(days=2)
next_day_name = days_ru[nextmorrow.weekday()]
nextmorrow = nextmorrow.strftime('%d.%m.%Y') + f' {next_day_name}'
if next_day_name == 'воскресенье':
    nextmorrow = datetime.now() + timedelta(days=3)
    next_day_name = days_ru[nextmorrow.weekday()]
    nextmorrow = nextmorrow.strftime('%d.%m.%Y') + f' {next_day_name}'
else:
    pass


# Окрытие файла, открытие первого листа
wb = load_workbook("check_table.xlsx")
shedule = wb.active

# Переменные для создания диапазона расписания
start_row = 0
end_row = 0
column_s = 0

# Поиск строки завтрашнего дня
for column in range(1, shedule.max_column+1):
    for row in range(1, shedule.max_row+1):
        if shedule.cell(column=column, row=row).value == tomorrow:
            start_row = shedule.cell(row=row, column=column).row

# Поиск строки ограничения(послезавтра)
for column in range(1, shedule.max_column+1):
    for row in range(1, shedule.max_row+1):
        if shedule.cell(column=column, row=row).value == nextmorrow:
            end_row = shedule.cell(row=row, column=column).row

# Поиск столбца группы
current_group = "Т-924/3 и Т-1125"
for column in range(1, shedule.max_column+1):
    for row in range(1, shedule.max_row+1):
        if shedule.cell(column=column, row=row).value == current_group:
            column_s = shedule.cell(row=row, column=column).column

# Составление расписания
para = 1
for i in range(start_row, end_row, 2):
    if shedule.cell(column=column_s, row=i).value == None:
        print(f"{para})---")
        para += 1
    else:
        print(f"{para}) {shedule.cell(column=column_s, row=i).value}".ljust(
            50), end="\t")
        print(f"{shedule.cell(column=column_s, row=i + 1).value}")
        print(f"{shedule.cell(column=column_s + 3, row=i).value}")
        para += 1
if os.path.exists("check_table.xlsx"):
    os.remove("check_table.xlsx")

input("Нажми на энетер, чтобы закрыть...")
